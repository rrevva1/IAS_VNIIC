#!/usr/bin/env python3
"""
Загрузка первых 20 строк из ОУ_тест.xlsx (лист АРМ) в БД ias_vniic, схема tech_accounting.
Парсинг строго по docs/РЕГЛАМЕНТ_ПАРСИНГА_ОУ_ТЕСТ.md.
"""
import os
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    import psycopg2
    from psycopg2.extras import execute_values
except ImportError as e:
    print("Требуются: pip install openpyxl psycopg2-binary")
    raise SystemExit(1) from e

# Конфиг (можно переопределить переменными окружения)
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = os.environ.get("OU_TEST_EXCEL", str(PROJECT_ROOT / "ОУ_тест.xlsx"))
DB_HOST = os.environ.get("PGHOST", "localhost")
DB_PORT = os.environ.get("PGPORT", "5432")
DB_NAME = os.environ.get("PGDATABASE", "ias_vniic")
DB_USER = os.environ.get("PGUSER", "postgres")
DB_PASSWORD = os.environ.get("PGPASSWORD", "12345")
ROWS_TO_LOAD = 20
SCHEMA = "tech_accounting"


def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    return v


def first_line_only(s):
    """По регламенту: при \\n в Пользователе брать первую строку."""
    if not s:
        return None
    s = str(s).strip()
    return s.split("\n")[0].strip() if "\n" in s else s


def inv_first_line(s):
    """Для № системн. блока при \\n — первая подстрока."""
    if not s:
        return None
    s = str(s).strip()
    return s.split("\n")[0].strip() if "\n" in s else s


def disk_value(s):
    """Диск: при \\n объединяем в один value_text (UNIQUE не даёт несколько записей на одну пару part/char)."""
    if not s:
        return None
    s = str(s).strip()
    return s.replace("\n", "; ") if "\n" in s else s


def monitor_value(s):
    """Монитор: аналогично — один value_text при нескольких через \\n."""
    if not s:
        return None
    s = str(s).strip()
    return s.replace("\n", "; ") if "\n" in s else s


def parse_date(val):
    """Год (int) -> 01.01.YYYY; DD.MM.YYYY -> date; иначе None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, (int, float)):
        y = int(val)
        if 1990 <= y <= 2030:
            return datetime(y, 1, 1).date()
        return None
    s = str(val).strip()
    if not s or "(" in s or "б/у" in s.lower():
        return None
    # DD.MM.YYYY
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            return None
    # только год
    if re.match(r"^\d{4}$", s):
        return datetime(int(s), 1, 1).date()
    return None


def location_name(val):
    """Помещение: число -> строка."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return str(int(val))
    return str(val).strip() or None


def load_excel_rows(path, sheet_name="АРМ", max_rows=20):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = list(ws.iter_rows(min_row=2, max_row=1 + max_rows, values_only=True))
    wb.close()
    return headers, rows


def get_col_index(headers, names):
    for name in names:
        for i, h in enumerate(headers):
            if h and name in h:
                return i
    return None


def main():
    if not Path(EXCEL_PATH).exists():
        print(f"Файл не найден: {EXCEL_PATH}")
        raise SystemExit(1)

    headers, rows = load_excel_rows(EXCEL_PATH, max_rows=ROWS_TO_LOAD)
    # Индексы столбцов (0-based)
    idx_user = get_col_index(headers, ["Пользователь"])
    idx_room = get_col_index(headers, ["Помещение"])
    idx_cpu = get_col_index(headers, ["ЦП"])
    idx_ram = get_col_index(headers, ["ОЗУ"])
    idx_disk = get_col_index(headers, ["Диск"])
    idx_sb = get_col_index(headers, ["Системный блок"])
    idx_inv_sb = get_col_index(headers, ["№ системн. блока"])
    idx_date_sb = get_col_index(headers, ["Дата закупки системного блока"])
    idx_monitor = get_col_index(headers, ["Монитор"])
    idx_inv_mon = get_col_index(headers, ["№ монитора"])
    idx_date_mon = get_col_index(headers, ["Дата закупки монитора"])
    idx_hostname = get_col_index(headers, ["Имя"])
    idx_ip = get_col_index(headers, ["IP адрес"])
    idx_os = get_col_index(headers, ["ОС"])
    idx_av = get_col_index(headers, ["Антивирус"])
    idx_other = get_col_index(headers, ["Другая техника"])

    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
    )
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(f"SET search_path TO {SCHEMA}")

    # Справочники: статус по умолчанию
    cur.execute("SELECT id FROM dic_equipment_status WHERE status_code = 'in_use' LIMIT 1")
    row = cur.fetchone()
    if not row:
        print("В БД нет статуса in_use. Выполните create_ias_vniic.sql")
        conn.rollback()
        conn.close()
        raise SystemExit(1)
    status_id = row[0]

    # Роль для новых пользователей
    cur.execute("SELECT id FROM roles WHERE role_code = 'user' LIMIT 1")
    r = cur.fetchone()
    role_id = r[0] if r else None

    def get_or_create_location(name):
        name = location_name(name)
        if not name:
            return None
        cur.execute("SELECT id FROM locations WHERE name = %s", (name,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute(
            "INSERT INTO locations (name, location_type) VALUES (%s, 'кабинет') RETURNING id",
            (name,),
        )
        return cur.fetchone()[0]

    def get_or_create_user(full_name):
        full_name = first_line_only(full_name)
        if not full_name:
            return None
        cur.execute("SELECT id FROM users WHERE full_name = %s AND is_deleted = FALSE", (full_name,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute(
            "INSERT INTO users (full_name) VALUES (%s) RETURNING id",
            (full_name,),
        )
        uid = cur.fetchone()[0]
        if role_id:
            cur.execute(
                "INSERT INTO user_roles (user_id, role_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                (uid, role_id),
            )
        return uid

    # Справочники частей и характеристик (создать при отсутствии)
    parts = {"ЦП": None, "ОЗУ": None, "Накопитель": None, "Монитор": None, "ПК": None}
    chars = {"Модель": None, "Объём": None, "№ монитора": None, "Имя ПК": None, "IP адрес": None, "ОС": None, "Антивирус": None}
    for name in parts:
        cur.execute("SELECT id FROM spr_parts WHERE name = %s", (name,))
        r = cur.fetchone()
        if r:
            parts[name] = r[0]
        else:
            cur.execute("INSERT INTO spr_parts (name) VALUES (%s) RETURNING id", (name,))
            parts[name] = cur.fetchone()[0]
    for name in chars:
        cur.execute("SELECT id FROM spr_chars WHERE name = %s", (name,))
        r = cur.fetchone()
        if r:
            chars[name] = r[0]
        else:
            cur.execute("INSERT INTO spr_chars (name) VALUES (%s) RETURNING id", (name,))
            chars[name] = cur.fetchone()[0]

    loaded = 0
    errors = []
    seen_inv = set()  # для уникальности инв. номера в рамках загрузки (дубликаты в файле)

    for row_num, row in enumerate(rows, start=2):
        if row_num - 2 >= ROWS_TO_LOAD:
            break
        inv_sb = inv_first_line(norm(row[idx_inv_sb]) if idx_inv_sb is not None and idx_inv_sb < len(row) else None)
        inv_mon = norm(row[idx_inv_mon]) if idx_inv_mon is not None and idx_inv_mon < len(row) else None
        has_sb = bool(inv_sb or (idx_sb is not None and idx_sb < len(row) and norm(row[idx_sb])))
        if has_sb and not inv_sb:
            inv_sb = inv_mon or f"МЦ.04-ROW-{row_num}"
        if not has_sb:
            inv_sb = inv_first_line(inv_mon) if inv_mon else f"МЦ.04-MON-{row_num}"
        if not inv_sb:
            errors.append(f"Строка {row_num}: нет инв. номера, пропуск")
            continue
        # Дубликаты инв. номера в файле: делаем уникальным суффиксом (регламент: обработка дубликатов)
        if inv_sb in seen_inv:
            inv_sb = f"{inv_sb}-строка{row_num}"
        seen_inv.add(inv_sb)

        loc_name = location_name(row[idx_room] if idx_room is not None and idx_room < len(row) else None)
        if not loc_name:
            errors.append(f"Строка {row_num}: нет помещения, пропуск")
            continue
        location_id = get_or_create_location(loc_name)
        if not location_id:
            continue

        user_id = None
        if idx_user is not None and idx_user < len(row):
            user_id = get_or_create_user(row[idx_user])

        name_equip = None
        equipment_type = "Системный блок"
        if idx_sb is not None and idx_sb < len(row) and norm(row[idx_sb]):
            name_equip = norm(row[idx_sb])
            if "моноблок" in (name_equip or "").lower() or "ноутбук" in (name_equip or "").lower():
                equipment_type = "Моноблок" if "моноблок" in name_equip.lower() else "Ноутбук"
        if not name_equip and idx_monitor is not None and idx_monitor < len(row) and norm(row[idx_monitor]):
            name_equip = monitor_value(row[idx_monitor]) or "Монитор"
            equipment_type = "Монитор"
        if not name_equip:
            name_equip = inv_sb
        if name_equip and "\n" in name_equip:
            name_equip = name_equip.replace("\n", "; ")

        purchase_date = None
        if idx_date_sb is not None and idx_date_sb < len(row):
            purchase_date = parse_date(row[idx_date_sb])

        desc_parts = []
        if idx_other is not None and idx_other < len(row) and norm(row[idx_other]):
            desc_parts.append(norm(row[idx_other]))

        try:
            cur.execute(
                """INSERT INTO equipment (inventory_number, name, equipment_type, status_id, responsible_user_id, location_id, purchase_date, description)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id""",
                (
                    inv_sb,
                    name_equip[:200],
                    equipment_type[:100],
                    status_id,
                    user_id,
                    location_id,
                    purchase_date,
                    "; ".join(desc_parts) if desc_parts else None,
                ),
            )
            equip_id = cur.fetchone()[0]

            # part_char_values (по регламенту; одна запись на (equipment_id, part_id, char_id))
            def add_pcv(part_name, char_name, value):
                if value is None or not str(value).strip():
                    return
                p_id = parts.get(part_name)
                c_id = chars.get(char_name)
                if p_id is None or c_id is None:
                    return
                cur.execute(
                    """INSERT INTO part_char_values (equipment_id, part_id, char_id, value_text)
                       VALUES (%s, %s, %s, %s)
                       ON CONFLICT (equipment_id, part_id, char_id) DO UPDATE SET value_text = EXCLUDED.value_text""",
                    (equip_id, p_id, c_id, str(value).strip()[:5000]),
                )

            if idx_cpu is not None and idx_cpu < len(row) and norm(row[idx_cpu]):
                add_pcv("ЦП", "Модель", norm(row[idx_cpu]))
            if idx_ram is not None and idx_ram < len(row) and norm(row[idx_ram]):
                add_pcv("ОЗУ", "Объём", norm(row[idx_ram]))
            if idx_disk is not None and idx_disk < len(row):
                add_pcv("Накопитель", "Модель", disk_value(row[idx_disk]))
            if idx_monitor is not None and idx_monitor < len(row) and norm(row[idx_monitor]):
                add_pcv("Монитор", "Модель", monitor_value(row[idx_monitor]))
            if idx_inv_mon is not None and idx_inv_mon < len(row) and norm(row[idx_inv_mon]):
                add_pcv("Монитор", "№ монитора", norm(row[idx_inv_mon]))
            if idx_hostname is not None and idx_hostname < len(row) and norm(row[idx_hostname]):
                add_pcv("ПК", "Имя ПК", norm(row[idx_hostname]))
            if idx_ip is not None and idx_ip < len(row) and norm(row[idx_ip]):
                add_pcv("ПК", "IP адрес", norm(row[idx_ip]))
            if idx_os is not None and idx_os < len(row) and norm(row[idx_os]):
                add_pcv("ПК", "ОС", norm(row[idx_os]))
            if idx_av is not None and idx_av < len(row) and norm(row[idx_av]):
                add_pcv("ПК", "Антивирус", norm(row[idx_av]))

            loaded += 1
        except psycopg2.IntegrityError as e:
            errors.append(f"Строка {row_num}: {e}")
            conn.rollback()
            conn.close()
            raise

    conn.commit()
    cur.close()
    conn.close()

    print(f"Загружено записей equipment: {loaded} из {ROWS_TO_LOAD}")
    if errors:
        for e in errors:
            print("  ", e)


if __name__ == "__main__":
    main()
