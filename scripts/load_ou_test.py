#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Загрузка данных из ОУ_тест.xlsx (лист АРМ) в БД ias_vniic, схема tech_accounting.
Правила парсинга: docs/РЕГЛАМЕНТ_ПАРСИНГА_ОУ_ТЕСТ.md.
Загружаются первые N записей (по умолчанию 50).

Использование:
  pip install -r scripts/requirements-load.txt
  python scripts/load_ou_test.py --file "d:/Projects/TZ/ОУ_тест.xlsx" --limit 50

  Параметры БД можно задать переменными PGHOST, PGPORT, PGDATABASE, PGUSER, PGPASSWORD
  или ключами --db-host, --db-name, --db-user, --db-password.
  Файл по умолчанию ищется в текущей директории и в корне проекта (родитель scripts/).
"""

import argparse
import logging
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal

try:
    import openpyxl
except ImportError:
    print("Установите openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)
try:
    import psycopg2
    from psycopg2.extras import execute_values
except ImportError:
    print("Установите psycopg2-binary: pip install psycopg2-binary", file=sys.stderr)
    sys.exit(1)
try:
    from dateutil import parser as date_parser
except ImportError:
    date_parser = None

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# Маппинг столбцов листа АРМ (индекс 0-based или по заголовку)
COL_USER = 0          # Пользователь
COL_OTDEL = 1         # Отдел — не загружаем
COL_ROOM = 2          # Помещение
COL_CPU = 3           # ЦП
COL_RAM = 4           # ОЗУ
COL_DISK = 5          # Диск
COL_SB = 6            # Системный блок
COL_BIOS_PASS = 7     # Пароль на BIOS — не загружаем
COL_INV_SB = 8        # № системн. блока
COL_DATE_SB = 9       # Дата закупки системного блока
COL_MONITOR = 10      # Монитор
COL_INV_MON = 11      # № монитора
COL_DATE_MON = 12     # Дата закупки монитора
COL_NAME_PC = 13      # Имя
COL_IP = 14           # IP адрес (в файле может быть пробел перед "IP")
COL_OS = 15           # ОС
COL_AV = 16           # Антивирус
COL_OTHER = 17        # Другая техника

# Статусы-не-отделы (регламент): не грузим в department
STATUS_VALUES = {
    "карточка оформлена", "карточка не нужна", "оформить карточку",
    "где мониторы?", "временно до 31.07.2025", "мол данилкин", ",",
}

# Типы частей и характеристик для part_char_values
PARTS_AND_CHARS = [
    ("ЦП", "Модель"),
    ("ОЗУ", "Объём"),
    ("Накопитель", "Наименование"),
    ("Монитор", "Модель"),
    ("Монитор", "№ монитора"),
    ("Монитор", "Дата закупки монитора"),
    ("ПК", "Имя ПК"),
    ("ПК", "IP адрес"),
    ("ПК", "ОС"),
    ("ПК", "Антивирус"),
]


def cell_str(ws, row, col):
    """Значение ячейки как строка; пусто → пустая строка."""
    v = ws.cell(row=row, column=col + 1).value
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    return str(v).strip()


def first_line(s):
    """При неатомарности (\\n) — первая строка (регламент: Пользователь, № системн. блока и т.д.)."""
    if not s:
        return ""
    return s.split("\n")[0].strip()


def split_by_newline(s):
    """Разбить по \\n; пустые подстроки отфильтровать."""
    if not s:
        return []
    return [x.strip() for x in s.split("\n") if x.strip()]


def parse_date(val):
    """Парсинг даты: год (int), DD.MM.YYYY, datetime, иначе None."""
    if val is None or val == "":
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, (int, float)):
        try:
            y = int(val)
            if 1990 <= y <= 2030:
                return date(y, 1, 1)
        except (ValueError, TypeError):
            pass
        return None
    s = str(val).strip()
    if not s:
        return None
    # Текст типа "2022 (б/у)", "Информтехника" — не парсить
    if re.search(r"\(б/у\)|информтехника", s, re.I):
        return None
    # DD.MM.YYYY
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except (ValueError, TypeError):
            pass
    if date_parser:
        try:
            dt = date_parser.parse(s, dayfirst=True)
            return dt.date() if hasattr(dt, "date") else dt
        except Exception:
            pass
    return None


def normalize_room(room):
    """Помещение: число → строка '103', '105'; текст как есть."""
    if room is None or str(room).strip() == "":
        return ""
    s = str(room).strip()
    try:
        n = int(Decimal(s.replace(",", ".")))
        return str(n)
    except Exception:
        return s


def equipment_type_from_name(name):
    """Вывести equipment_type из наименования СБ/монитора."""
    if not name:
        return "Системный блок"
    n = name.lower()
    if "моноблок" in n or "monoblock" in n:
        return "Моноблок"
    if "ноутбук" in n or "laptop" in n:
        return "Ноутбук"
    if "монитор" in n or "monitor" in n:
        return "Монитор"
    return "Системный блок"


def read_arm_sheet(ws, max_rows=None):
    """Читает лист АРМ: первая строка — заголовки, далее данные. Возвращает список dict по столбцам."""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=18, values_only=False))
    if not rows:
        return []
    # Проверка заголовков по первой строке (опционально)
    data_start = 2  # строка 2 в Excel = первая строка данных
    result = []
    for r in range(1, len(rows)):
        row_idx = r + 1
        row_cells = rows[r]
        if max_rows and len(result) >= max_rows:
            break
        row_dict = {}
        for col in range(18):
            val = row_cells[col].value if col < len(row_cells) else None
            if val is not None and isinstance(val, (datetime, date)):
                val = val.isoformat() if hasattr(val, "isoformat") else str(val)
            elif val is not None:
                val = str(val).strip() if isinstance(val, str) else val
            row_dict[col] = val
        result.append((row_idx + 1, row_dict))  # Excel row 1-based
    return result


def extract_row(excel_row_index, row_dict):
    """Из одной строки листа извлекает нормализованные поля и список part_char_values по регламенту."""
    def c(col):
        v = row_dict.get(col)
        if v is None:
            return ""
        return str(v).strip() if isinstance(v, str) else str(v).strip()

    user_raw = c(COL_USER)
    room_raw = c(COL_ROOM)
    cpu = c(COL_CPU)
    ram = c(COL_RAM)
    disk = c(COL_DISK)
    sb = c(COL_SB)
    inv_sb = c(COL_INV_SB)
    date_sb_raw = row_dict.get(COL_DATE_SB)
    monitor = c(COL_MONITOR)
    inv_mon = c(COL_INV_MON)
    date_mon_raw = row_dict.get(COL_DATE_MON)
    name_pc = c(COL_NAME_PC)
    ip = c(COL_IP)
    os_ = c(COL_OS)
    av = c(COL_AV)
    other = c(COL_OTHER)

    # Пользователь: первая строка при \n
    full_name = first_line(user_raw) if user_raw else ""
    location_name = normalize_room(room_raw) if room_raw else ""
    purchase_date = parse_date(date_sb_raw)

    only_monitor = not (cpu or ram or disk or sb or inv_sb)
    if only_monitor:
        inv_number = first_line(inv_mon) if inv_mon else ""
        name = first_line(monitor) if monitor else "Монитор"
        eq_type = "Монитор"
    else:
        inv_number = first_line(inv_sb) if inv_sb else ""
        name = first_line(sb) if sb else "Системный блок"
        eq_type = equipment_type_from_name(sb)

    description_parts = []
    if date_sb_raw and not purchase_date and str(date_sb_raw).strip():
        description_parts.append("Дата закупки СБ: " + str(date_sb_raw).strip())
    if other:
        description_parts.append("Другая техника: " + other)
    description = " ".join(description_parts) if description_parts else None

    chars = []  # список (part_name, char_name, value_text)

    # В БД UNIQUE(equipment_id, part_id, char_id) — одна запись на пару (часть, характеристика); несколько значений объединяем через \n
    if cpu:
        chars.append(("ЦП", "Модель", cpu))
    if ram:
        chars.append(("ОЗУ", "Объём", ram))
    disk_parts = [p for p in split_by_newline(disk) if p]
    if disk_parts:
        chars.append(("Накопитель", "Наименование", "\n".join(disk_parts)))
    monitor_parts = [p for p in split_by_newline(monitor) if p]
    if monitor_parts:
        chars.append(("Монитор", "Модель", "\n".join(monitor_parts)))
    inv_mon_parts = [p for p in split_by_newline(inv_mon) if p]
    if inv_mon_parts:
        chars.append(("Монитор", "№ монитора", "\n".join(inv_mon_parts)))
    if date_mon_raw:
        mon_date = parse_date(date_mon_raw)
        if mon_date:
            chars.append(("Монитор", "Дата закупки монитора", str(mon_date)))
        else:
            chars.append(("Монитор", "Дата закупки монитора", str(date_mon_raw).strip()))
    if name_pc:
        chars.append(("ПК", "Имя ПК", name_pc))
    if ip:
        chars.append(("ПК", "IP адрес", ip))
    if os_:
        chars.append(("ПК", "ОС", os_))
    if av:
        chars.append(("ПК", "Антивирус", av))

    return {
        "excel_row": excel_row_index,
        "full_name": full_name,
        "location_name": location_name,
        "inventory_number": inv_number,
        "name": name or "—",
        "equipment_type": eq_type,
        "purchase_date": purchase_date,
        "description": description,
        "part_char_values": chars,
    }


def ensure_schema(conn):
    conn.cursor().execute("SET search_path TO tech_accounting;")


def get_or_create_location(conn, name, cur):
    """Вернуть id локации; при отсутствии создать (location_type = 'кабинет')."""
    name = (name or "").strip()
    if not name:
        return None
    cur.execute(
        "SELECT id FROM tech_accounting.locations WHERE name = %s",
        (name,),
    )
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(
        """INSERT INTO tech_accounting.locations (name, location_type)
           VALUES (%s, 'кабинет') ON CONFLICT (name) DO NOTHING""",
        (name,),
    )
    cur.execute("SELECT id FROM tech_accounting.locations WHERE name = %s", (name,))
    row = cur.fetchone()
    return row[0] if row else None


def get_or_create_user(conn, full_name, cur):
    """Вернуть id пользователя; при отсутствии создать с minimal данными."""
    full_name = (full_name or "").strip()
    if not full_name:
        return None
    cur.execute(
        "SELECT id FROM tech_accounting.users WHERE full_name = %s AND is_deleted = FALSE",
        (full_name,),
    )
    row = cur.fetchone()
    if row:
        return row[0]
    username = "import_" + re.sub(r"\W+", "_", full_name)[:80]
    cur.execute(
        """INSERT INTO tech_accounting.users (full_name, username)
           VALUES (%s, %s)
           ON CONFLICT (username) DO UPDATE SET full_name = EXCLUDED.full_name
           RETURNING id""",
        (full_name, username),
    )
    r = cur.fetchone()
    if r:
        return r[0]
    cur.execute("SELECT id FROM tech_accounting.users WHERE full_name = %s", (full_name,))
    row = cur.fetchone()
    return row[0] if row else None


def ensure_parts_and_chars(conn, cur):
    """Заполнить spr_parts и spr_chars нужными именами; вернуть словари name -> id."""
    parts = {}
    chars = {}
    for part_name, char_name in PARTS_AND_CHARS:
        if part_name not in parts:
            cur.execute(
                """INSERT INTO tech_accounting.spr_parts (name) VALUES (%s)
                   ON CONFLICT (name) DO NOTHING""",
                (part_name,),
            )
            cur.execute("SELECT id FROM tech_accounting.spr_parts WHERE name = %s", (part_name,))
            parts[part_name] = cur.fetchone()[0]
        if char_name not in chars:
            cur.execute(
                """INSERT INTO tech_accounting.spr_chars (name) VALUES (%s)
                   ON CONFLICT (name) DO NOTHING""",
                (char_name,),
            )
            cur.execute("SELECT id FROM tech_accounting.spr_chars WHERE name = %s", (char_name,))
            chars[char_name] = cur.fetchone()[0]
    return parts, chars


def get_status_id_in_use(cur):
    cur.execute(
        "SELECT id FROM tech_accounting.dic_equipment_status WHERE status_code = %s",
        ("in_use",),
    )
    row = cur.fetchone()
    if not row:
        raise RuntimeError("В БД нет статуса dic_equipment_status.status_code = 'in_use'")
    return row[0]


def load_one_row(conn, cur, rec, location_cache, user_cache, parts, chars, status_id):
    """Вставить одну запись equipment и связанные part_char_values. При дубликате inventory_number — пропуск с ошибкой."""
    loc_id = None
    if rec["location_name"]:
        if rec["location_name"] not in location_cache:
            location_cache[rec["location_name"]] = get_or_create_location(conn, rec["location_name"], cur)
        loc_id = location_cache[rec["location_name"]]
    if not loc_id:
        return False, "Пустое или не создано помещение"

    user_id = None
    if rec["full_name"]:
        if rec["full_name"] not in user_cache:
            user_cache[rec["full_name"]] = get_or_create_user(conn, rec["full_name"], cur)
        user_id = user_cache[rec["full_name"]]

    inv = (rec["inventory_number"] or "").strip()
    if not inv:
        return False, "Пустой инв. номер"

    cur.execute(
        "SELECT id FROM tech_accounting.equipment WHERE inventory_number = %s",
        (inv,),
    )
    if cur.fetchone():
        return False, f"Дубликат inventory_number: {inv}"

    cur.execute(
        """INSERT INTO tech_accounting.equipment
           (inventory_number, name, equipment_type, status_id, responsible_user_id, location_id, purchase_date, description)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
           RETURNING id""",
        (
            inv,
            rec["name"] or "—",
            rec["equipment_type"] or "Системный блок",
            status_id,
            user_id,
            loc_id,
            rec["purchase_date"],
            rec["description"],
        ),
    )
    eq_id = cur.fetchone()[0]

    for part_name, char_name, value_text in rec["part_char_values"]:
        pid = parts.get(part_name)
        cid = chars.get(char_name)
        if not pid or not cid or not value_text:
            continue
        cur.execute(
            """INSERT INTO tech_accounting.part_char_values (equipment_id, part_id, char_id, value_text, source)
               VALUES (%s, %s, %s, %s, 'import')
               ON CONFLICT (equipment_id, part_id, char_id) DO UPDATE SET value_text = EXCLUDED.value_text, source = 'import'""",
            (eq_id, pid, cid, value_text[:10000]),
        )
    return True, eq_id


def main():
    ap = argparse.ArgumentParser(description="Загрузка ОУ_тест.xlsx (лист АРМ) в ias_vniic")
    ap.add_argument("--file", default="ОУ_тест.xlsx", help="Путь к файлу ОУ_тест.xlsx")
    ap.add_argument("--limit", type=int, default=50, help="Максимум записей для загрузки (по умолчанию 50)")
    ap.add_argument("--db-host", default=os.environ.get("PGHOST", "localhost"))
    ap.add_argument("--db-port", default=os.environ.get("PGPORT", "5432"))
    ap.add_argument("--db-name", default=os.environ.get("PGDATABASE", "ias_vniic"))
    ap.add_argument("--db-user", default=os.environ.get("PGUSER", "postgres"))
    ap.add_argument("--db-password", default=os.environ.get("PGPASSWORD", ""))
    args = ap.parse_args()

    path = args.file
    if not os.path.isfile(path):
        # Попробовать в корне проекта (родитель scripts/)
        root_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "ОУ_тест.xlsx")
        if os.path.isfile(root_file):
            path = root_file
        else:
            log.error("Файл не найден: %s", args.file)
            sys.exit(1)

    log.info("Чтение листа АРМ из %s, лимит %d записей", path, args.limit)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    if "АРМ" not in wb.sheetnames:
        log.error("Лист 'АРМ' не найден. Листы: %s", wb.sheetnames)
        sys.exit(1)
    ws = wb["АРМ"]

    rows_with_dicts = []
    for r in range(2, min(ws.max_row + 1, 2 + args.limit)):
        row_dict = {}
        for col in range(18):
            row_dict[col] = ws.cell(row=r, column=col + 1).value
        rows_with_dicts.append((r, row_dict))

    conn = psycopg2.connect(
        host=args.db_host,
        port=args.db_port,
        dbname=args.db_name,
        user=args.db_user,
        password=args.db_password or None,
    )
    conn.autocommit = False
    cur = conn.cursor()
    ensure_schema(conn)

    status_id = get_status_id_in_use(cur)
    ensure_parts_and_chars(conn, cur)
    parts, chars = {}, {}
    cur.execute("SELECT name, id FROM tech_accounting.spr_parts")
    for name, id_ in cur.fetchall():
        parts[name] = id_
    cur.execute("SELECT name, id FROM tech_accounting.spr_chars")
    for name, id_ in cur.fetchall():
        chars[name] = id_

    # Запуск импорта
    cur.execute(
        """INSERT INTO tech_accounting.import_runs (source_type, source_name, total_rows, run_status)
           VALUES ('ou_test_xlsx', %s, %s, 'running')
           RETURNING id""",
        (os.path.basename(path), len(rows_with_dicts)),
    )
    run_id = cur.fetchone()[0]
    conn.commit()

    success = 0
    errors = []
    location_cache = {}
    user_cache = {}

    for excel_row, row_dict in rows_with_dicts:
        rec = extract_row(excel_row, row_dict)
        try:
            ok, msg = load_one_row(conn, cur, rec, location_cache, user_cache, parts, chars, status_id)
            if ok:
                success += 1
            else:
                errors.append((excel_row, msg))
                cur.execute(
                    """INSERT INTO tech_accounting.import_errors (import_run_id, row_number, error_message)
                       VALUES (%s, %s, %s)""",
                    (run_id, excel_row, msg),
                )
        except Exception as e:
            errors.append((excel_row, str(e)))
            cur.execute(
                """INSERT INTO tech_accounting.import_errors (import_run_id, row_number, error_message)
                   VALUES (%s, %s, %s)""",
                (run_id, excel_row, str(e)),
            )

    run_status = "success" if not errors else "partial" if success else "error"
    cur.execute(
        """UPDATE tech_accounting.import_runs
           SET finished_at = CURRENT_TIMESTAMP, success_rows = %s, error_rows = %s, run_status = %s
           WHERE id = %s""",
        (success, len(errors), run_status, run_id),
    )
    conn.commit()
    cur.close()
    conn.close()

    log.info("Импорт завершён: run_id=%s, успешно=%s, ошибок=%s, статус=%s", run_id, success, len(errors), run_status)
    for row, err in errors[:20]:
        log.warning("Строка %s: %s", row, err)
    if len(errors) > 20:
        log.warning("... и ещё %s ошибок", len(errors) - 20)
    wb.close()
    return 0 if run_status != "error" else 1


if __name__ == "__main__":
    sys.exit(main())
